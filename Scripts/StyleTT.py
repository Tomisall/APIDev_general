from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill

def StyleTT(outputName):
	wb_style_prod = load_workbook(outputName) 
	#Accessing Product Informaiton Sheet
	sheetA = wb_style_prod['2d.p. %area']
	sheetB = wb_style_prod['2d.p. area']
	sheetC = wb_style_prod['3d.p. %area']
	sheetD = wb_style_prod['3d.p. area']
	sheetE = wb_style_prod['Simple Output']
	#Setting background heading row only
	sheetA.insert_cols(2)
	sheetA.move_range("A1:A3", cols=1)
	sheetA.merge_cells('A1:A3')
	sheetA['A4'].value='Reference'
	sheetA['B4'].value='Comment'
	sheetB.insert_cols(2)
	sheetB.move_range("A1:A3", cols=1)
	sheetB.merge_cells('A1:A3')
	sheetB['A4'].value='Reference'
	sheetB['B4'].value='Comment'
	sheetC.insert_cols(2)
	sheetC.move_range("A1:A3", cols=1)
	sheetC.merge_cells('A1:A3')
	sheetC['A4'].value='Reference'
	sheetC['B4'].value='Comment'
	sheetD.insert_cols(2)
	sheetD.move_range("A1:A3", cols=1)
	sheetD.merge_cells('A1:A3')
	sheetD['A4'].value='Reference'
	sheetD['B4'].value='Comment'
	sheetE.insert_cols(2)
	sheetE.move_range("A1:A3", cols=1)
	sheetE.merge_cells('A1:A3')
	sheetE['A4'].value='Reference'
	sheetE['B4'].value='Comment'
	for rows in sheetA.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetB.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetC.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetD.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetE.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetA.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
	for rows in sheetB.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
	for rows in sheetC.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
			cell.number_format = "0.000"
	for rows in sheetD.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
			cell.number_format = "0.000"
	for rows in sheetE.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
	sheetA.column_dimensions['A'].width = 30
	sheetA.column_dimensions['B'].width = 12
	sheetB.column_dimensions['A'].width = 30
	sheetB.column_dimensions['B'].width = 12
	sheetC.column_dimensions['A'].width = 30
	sheetC.column_dimensions['B'].width = 12
	sheetD.column_dimensions['A'].width = 30
	sheetD.column_dimensions['B'].width = 12
	sheetE.column_dimensions['A'].width = 30
	sheetE.column_dimensions['B'].width = 12

	wb_style_prod.save(outputName)