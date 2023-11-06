import matplotlib.pyplot as plt
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
import warnings

warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)

wd = os.getcwd() + '\\Trend_CSVs\\'
print('Trending HPLC Integrals')

def trendHPLC(path,file):
	sheetDict = {}
	fullName = path+file
	xls = pd.ExcelFile(fullName)
	fullDict = pd.concat(pd.read_excel(fullName, sheet_name=None), ignore_index=True)
	rtONEList = fullDict.query("RRT==1")["RT [min]"]
	rtOneAVG = rtONEList.sum()/len(rtONEList)
	print(fullDict)
	print('\n')
	rrtIndex = fullDict["RRT"].values.tolist()
	rrtIndex = list(dict.fromkeys(rrtIndex))
	rrtIndex.sort()
	#print(rrtIndex)
	#print('\n')
	sampleIndex = fullDict["Sample Name"].values.tolist()
	sampleIndex = list(dict.fromkeys(sampleIndex))
	#print(sampleIndex)
	#print('\n')
	cmpnameIndex = fullDict["Compound Name"].values.tolist()
	cmpnameIndex = list(dict.fromkeys(rrtIndex))
	cmpdfIndex = []
	for rrtValue in rrtIndex:
		namedfValue = fullDict.loc[(fullDict['RRT'] == rrtValue), 'Compound Name'].dropna()
		nameValue = namedfValue.array
		#print(nameValue)
		if len(nameValue) == 1:
			cmpdfIndex.append(nameValue[0])
		elif len(nameValue) == 0:
			cmpdfIndex.append(' ')
		else:
			condnesdArry = list(dict.fromkeys(nameValue))
			#print(condnesdArry)
			cmpdfIndex.append(condnesdArry[0])

		
	#print(cmpdfIndex)
	cmpdfIndex.append(' ')
	areaIndex = []
	#print(type(rrtIndex))
	dfIndex = rrtIndex.copy()
	dfIndex.append('Sum')
	trendedDF = pd.DataFrame(dfIndex)
	trendedDF['Compound'] = cmpdfIndex 
	counteri = 0
	for sampleName in sampleIndex:
		tempList = []
		sumList = []
		for rrtValue in rrtIndex:
			areadfValue = fullDict.loc[(fullDict['Sample Name'] == sampleName) & (fullDict['RRT'] == rrtValue), 'Area']
			areaValue = areadfValue.array
			if len(areaValue) > 0:
				tempList.append(areaValue[0])
				sumList.append(areaValue[0])
			else:
				tempList.append('-')

		sumValue = sum(sumList)
		tempList.append(sumValue) 
		trendedDF[sampleName] = tempList

	#print(rawtrendDict)
	#print(areaIndex)
	trendedDF = trendedDF.T
	trendedDF.columns = trendedDF.iloc[0]
	#print(trendedDF)
	#print('\n')
	#print(len(rrtIndex))
	


	combineNames = cmpdfIndex.copy()
	combineNames = list(dict.fromkeys(combineNames))
	combineNames.remove(' ')
	#print(combineNames)
	#print('\n\n')
	headingsList = trendedDF.columns.tolist()
	for cmpdName in combineNames:
		testvalue = trendedDF.columns[trendedDF.isin([cmpdName]).any()] 
		#testvalue = trendedDF.loc[(trendedDF[0] == cmpdName)]
		#print(testvalue)
		#print('\n')
		#print(len(testvalue))
		#print('\n')
		avgRRT = round(sum(testvalue)/len(testvalue), 3)
		cmpdNewCol = [avgRRT, cmpdName]
		take2CMPDlist = []
		for rrtHeader in testvalue:
			headingsList.remove(rrtHeader)
		santyDF = trendedDF.drop(testvalue, axis=1)
		for index,row in trendedDF.iterrows():
			listforLen = []
			#smalldf = trendedDF[trendedDF.isin(testvalue)][row]
			#print(smalldf)
			currentRowValues = []
			for rrtHeader in testvalue:	
				cell = row[rrtHeader]
				currentRowValues.append(cell)
				for cmpdVals in currentRowValues:
					if cell != '-' and cell !=cmpdName and cell !=rrtHeader:
						cmpdNewCol.append(cell)
					else:
						listforLen.append(cell)
			#print(currentRowValues)
			minimisedList = list(dict.fromkeys(currentRowValues))
			if minimisedList.count('-') > 0:
				minimisedList.remove('-')
			if len(minimisedList) == 0:
				minimisedList.append('-')
			if len(minimisedList) > 1:
				take2CMPDlist.append(minimisedList)
			else:
				take2CMPDlist.extend(minimisedList)
			#print(len(minimisedList))

		headingsList.append(avgRRT)
		take2CMPDlist[0] = avgRRT
		#take2CMPDlist.replace('','-')		
		#print(take2CMPDlist)
		#print(len(cmpdNewCol))
		#print('\n')
		santyDF[avgRRT] = take2CMPDlist
				
		print(headingsList)
		headingsList.remove('Sum')
		headingsList = list(dict.fromkeys(headingsList))
		headingsList = [x for x in headingsList if str(x) != 'nan']
		headingsList.sort()
		headingsList.append('Sum')
		print(headingsList)

		#headingsList.remove(1.189)
		#headingsList.append(1.164)
		#headingsList.append(1.189)

	print(headingsList)	
	#headingsList.remove(1.164)
	#headingsList.remove(1.189)
	santyDF = santyDF[headingsList]
	#santyDF.reindex(columns=headingsList)
	#print('\n\n')
	#print(santyDF)
	#print('\n')
	#print(santyDF[1])

	outputName = 'output_' + str(file)
	santyZeroDF = santyDF.replace('-', 0)
	print('\n\n')
	print(santyZeroDF)
	print('\n\n')
	listfornewRow = santyDF.iloc[0].tolist()
	fixA = list(dict.fromkeys(listfornewRow))
	listfornewRow2 = santyDF.iloc[1].tolist()
	#fixB = list(dict.fromkeys(listfornewRow2))
	#print(listfornewRow,listfornewRow2)
	#print(fixA)
	print('\n\n')
	#print(fixB)
	percentAreaDF = pd.DataFrame(listfornewRow2,listfornewRow).T
	#percentAreaDF = pd.DataFrame(listfornewRow2,fixA).T
	blankAreaDF = pd.DataFrame()
	#percentAreaDF.columns = percentAreaDF.iloc[0]
	print(percentAreaDF)
	print('\n\n')
	for column in santyZeroDF.columns:
		#print(column)
		#print('\n')
		#print(round(santyZeroDF[column][2:].astype(float) / santyDF['Sum'][2:].astype(float) * 100, 2))
		blankAreaDF[column] = round(santyZeroDF[column][2:].astype(float) / santyDF['Sum'][2:].astype(float) * 100, 2)
		#fixA = round(santyZeroDF[column][2:].astype(float) / santyDF['Sum'][2:].astype(float) * 100, 2)
		#fixB = list(dict.fromkeys(fixA))
		#blankAreaDF[column] = fixB.append('---')
	#print('\n\n\n\n')
	#print(blankAreaDF)
	for index,row in blankAreaDF.iterrows():
		percentAreaDF.loc[index] = row
			
	percentAreaDF.rename(index={0:'Compound'},inplace=True)
	print(percentAreaDF)
	prettyperAreaDF = percentAreaDF.replace(0, '-')
	#print(prettyperAreaDF)

	#print('\n\n')
	#print(listfornewRow)
	twodpRRTs = [ round(elem, 2) for elem in listfornewRow[:-1]]
	threedpRTs = [ round(elem * rtOneAVG, 3) for elem in listfornewRow[:-1]]
	threedpRTs.append('Sum')
	rtOneAVG = float(rtOneAVG)
	twodpAVGRTs = [round(i * rtOneAVG,2) for i in twodpRRTs]
	twodpAVGRTs = list(dict.fromkeys(twodpAVGRTs))
	twodpRRTs.append('')
	row1RRTstwo = list(dict.fromkeys(twodpRRTs))
	twodpAVGRTs.append('Sum')
	#print(twodpAVGRTs)
	#print(twodpRRTs)
	#print('\n\n')
	
	
	twodpDF = pd.DataFrame(percentAreaDF)
	twodpDF.columns = twodpRRTs
	twodpDF = twodpDF.groupby(axis=1, level=0).sum()
	listCMPDnew = twodpDF.iloc[0].values.tolist()
	#print(listCMPDnew)
	#print('\n\n')
	new_labels = pd.MultiIndex.from_arrays([twodpAVGRTs, row1RRTstwo, listCMPDnew], names=['RT', 'RRT', 'Compound'])
	twodpDF = twodpDF.set_axis(new_labels, axis=1).iloc[1:]
	prettytwodpDF = twodpDF.replace(0, '-')
	#print('\n\n')
	#print(prettytwodpDF)

	finalformDF = santyDF.replace('-', 0)
	areatwodpDF = finalformDF.tail(-1)
	areatwodpDF.columns = twodpRRTs
	#print(areatwodpDF)
	areatwodpDF = areatwodpDF.groupby(axis=1, level=0).sum()
	new_labelstwo = pd.MultiIndex.from_arrays([twodpAVGRTs, row1RRTstwo, listCMPDnew], names=['RT', 'RRT', 'Compound'])
	areatwodpDF = areatwodpDF.set_axis(new_labelstwo, axis=1).iloc[1:]
	prettytwodpAREADF = areatwodpDF.replace(0, '-')
	print(prettytwodpAREADF)
	#print('\n\n')
	clean3dpArea = finalformDF.tail(-1).replace(0, '-')
	threedpRRTs = clean3dpArea.columns.tolist()
	threedpRRTs[-1] = ''
	listCMPDLong = clean3dpArea.iloc[0].values.tolist()
	new_labelsthree = pd.MultiIndex.from_arrays([threedpRTs, threedpRRTs, listCMPDLong], names=['RT', 'RRT', 'Compound'])
	clean3dpArea = clean3dpArea.set_axis(new_labelsthree, axis=1).iloc[1:]
	new_labelsfour = pd.MultiIndex.from_arrays([threedpRTs, threedpRRTs, listCMPDLong], names=['RT', 'RRT', 'Compound'])
	prettyperAreaDF = prettyperAreaDF.set_axis(new_labelsfour, axis=1).iloc[1:] 

	#Remove Static 'Sum' Column from %area sheets, now that check has been completed
	prettytwodpDF = prettytwodpDF.drop(columns=['Sum'])
	prettyperAreaDF = prettyperAreaDF.drop(columns=['Sum'])	

	simpleDF = pd.DataFrame(percentAreaDF)
	simpleDF = simpleDF.drop(columns=['Sum'])
	simptwodpRRTs = twodpRRTs[:-1]
	simpleDF.columns = simptwodpRRTs
	simpleDF = simpleDF.groupby(axis=1, level=0).sum()
	simpNameList = []
	simplefideList = []
	namedList = []
	namedonlynames = []
	fullsimpRRTList = []
	for name, values in prettytwodpDF.items(): 
		simpNameList.append(name[2])
		fullsimpRRTList.append(name[1])
		print(name)
		if name[2].isspace() or name[2] == '':
			simplefideList.append(name[1])
		else:
			namedList.append(name[1])
			namedonlynames.append(name[2])

	print('\n')
	print(simplefideList)
	dumbWorkAround = 0
	skipCheck = 0
	groupedAvgs = []
	for item in simplefideList:
		nextIndex = dumbWorkAround + 1
		if len(simplefideList) > nextIndex and skipCheck == 0:
			if simplefideList[nextIndex] - item < 0.02:
				mergeItem = (item + simplefideList[nextIndex])/2
				skipCheck = 1
			else:
				mergeItem = item
				skipCheck = 0
		else:
			mergeItem = item
			skipCheck = 0

		groupedAvgs.append(mergeItem)
		dumbWorkAround += 1

	print('\n')
	#print(groupedAvgs)
	formattedAvgs = [ '%.2f' % elem for elem in groupedAvgs ]
	print('\n')
	formattedAvgs = list(dict.fromkeys(formattedAvgs))
	print(formattedAvgs)

	dumbWorkAround = 0
	skipCheck = 0
	roundtwoAvgs = []
	for item in formattedAvgs:
		nextIndex = dumbWorkAround + 1
		if len(formattedAvgs) > nextIndex and skipCheck == 0:
			if float(formattedAvgs[nextIndex]) - float(item) < 0.02:
				mergeItem = (float(item) + float(formattedAvgs[nextIndex]))/2
				skipCheck = 1
			else:
				mergeItem = float(item)
				skipCheck = 0
		else:
			mergeItem = float(item)
			skipCheck = 0

		roundtwoAvgs.append(mergeItem)
		dumbWorkAround += 1
	print('\n')
	#print(roundtwoAvgs)
	formattedtwoAvgs = [ '%.2f' % elem for elem in roundtwoAvgs ]
	print('\n')
	formattedtwoAvgs = list(dict.fromkeys(formattedtwoAvgs))
	print(formattedtwoAvgs)

	dumbWorkAround = 0
	skipCheck = 0
	roundthreeAvgs = []
	for item in formattedtwoAvgs:
		nextIndex = dumbWorkAround + 1
		if len(formattedtwoAvgs) > nextIndex and skipCheck == 0:
			if float(formattedtwoAvgs[nextIndex]) - float(item) < 0.02:
				mergeItem = (float(item) + float(formattedtwoAvgs[nextIndex]))/2
				skipCheck = 1
			else:
				mergeItem = float(item)
				skipCheck = 0
		else:
			mergeItem = float(item)
			skipCheck = 0

		roundthreeAvgs.append(mergeItem)
		dumbWorkAround += 1
	print('\n')
	#print(roundthreeAvgs)
	formattedthreeAvgs = [ '%.2f' % elem for elem in roundthreeAvgs ]
	print('\n')
	formattedthreeAvgs = list(dict.fromkeys(formattedthreeAvgs))
	formattedthreeAvgs = [float(i) for i in formattedthreeAvgs]
	print(formattedthreeAvgs)


	dumbWorkAround = 0
	skipCheck = 0
	roundfourAvgs = []
	for item in formattedthreeAvgs:
		nextIndex = dumbWorkAround + 1
		previousIndex = dumbWorkAround - 1
		if len(formattedthreeAvgs) > nextIndex and skipCheck == 0:
			if float(formattedthreeAvgs[nextIndex]) - float(item) < 0.02:
				mergeItem = (float(item) + float(formattedthreeAvgs[nextIndex]))/2
				skipCheck = 1
			else:
				mergeItem = float(item)
				skipCheck = 0
		elif skipCheck == 1:
			mergeItem = float(roundfourAvgs[previousIndex])
			skipCheck = 0
		else:
			mergeItem = float(item)
			skipCheck = 0

		roundfourAvgs.append(mergeItem)
		dumbWorkAround += 1
	print('\n')
	#print(roundthreeAvgs)
	formattedfourAvgs = [ '%.2f' % elem for elem in roundfourAvgs ]
	print('\n')
	formattedfourAvgs = list(dict.fromkeys(formattedfourAvgs))
	formattedfourAvgs = [float(i) for i in formattedfourAvgs]
	print(formattedfourAvgs)

	for extractCMPDrrt in namedList:
		formattedfourAvgs.append(extractCMPDrrt)
	formattedfourAvgs.sort()
	print('\n\n')
	print(formattedfourAvgs)
	print('\n')

	newRow = []
	for orginalValue in fullsimpRRTList:
		closestValue = min(formattedfourAvgs, key=lambda x: abs(orginalValue - x))
		newRow.append(closestValue)
	print(newRow)

	newRTs = [round(i * rtOneAVG,2) for i in newRow]
	newnewRTs = [round(i * rtOneAVG,2) for i in formattedfourAvgs]
	newCMPDnamesList = []
	for foundRRT in formattedfourAvgs:
		if foundRRT in namedList:
			print('\n')
			print(foundRRT)
			foundNameIndex = namedList.index(foundRRT)
			foundName = namedonlynames[foundNameIndex]
			print(foundName)
			newCMPDnamesList.append(foundName)
		else:
			newCMPDnamesList.append(' ')			

	print(newRTs)
	print(simpNameList)
	print('\n')
	print(len(newRow))
	print(len(newRTs))
	print(len(simpNameList))
	print('\n\n')
	newnew_labels = pd.MultiIndex.from_arrays([newRTs, newRow, simpNameList], names=['RT', 'RRT', 'Compound'])
	simpleDF = simpleDF.set_axis(newnew_labels, axis=1).iloc[1:]
	print(simpleDF)
	print('\n')
	simpleDF = simpleDF.groupby(axis=1, level='RRT', group_keys=True).sum()
	simpleDF = simpleDF.replace(0, '-')
	newnewnew_labels = pd.MultiIndex.from_arrays([newnewRTs, formattedfourAvgs, newCMPDnamesList], names=['RT', 'RRT', 'Compound'])
	simpleDF = simpleDF.set_axis(newnewnew_labels, axis=1).iloc[0:]
	#simpleDF = simpleDF.reset_index()
	print('\n')
	print(simpleDF)

	with pd.ExcelWriter(outputName) as writer: 
		prettytwodpDF.to_excel(writer, sheet_name='2d.p. %area', index=True)
		prettytwodpAREADF.to_excel(writer, sheet_name='2d.p. area', index=True)
		prettyperAreaDF.to_excel(writer, sheet_name='3d.p. %area', index=True)
		clean3dpArea.to_excel(writer, sheet_name='3d.p. area', index=True)
		simpleDF.to_excel(writer, sheet_name='Simple Output', index=True)

	wb_style_prod = load_workbook(outputName) 
	#Accessing Product Informaiton Sheet
	sheetA = wb_style_prod['2d.p. %area']
	sheetB = wb_style_prod['2d.p. area']
	sheetC = wb_style_prod['3d.p. %area']
	sheetD = wb_style_prod['3d.p. area']
	sheetE = wb_style_prod['Simple Output']
	#Setting background heading row only
	sheetA.insert_cols(2)
	sheetA.move_range("A1:A3", cols=1)
	sheetA.merge_cells('A1:A3')
	sheetA['A4'].value='Reference'
	sheetA['B4'].value='Comment'
	sheetB.insert_cols(2)
	sheetB.move_range("A1:A3", cols=1)
	sheetB.merge_cells('A1:A3')
	sheetB['A4'].value='Reference'
	sheetB['B4'].value='Comment'
	sheetC.insert_cols(2)
	sheetC.move_range("A1:A3", cols=1)
	sheetC.merge_cells('A1:A3')
	sheetC['A4'].value='Reference'
	sheetC['B4'].value='Comment'
	sheetD.insert_cols(2)
	sheetD.move_range("A1:A3", cols=1)
	sheetD.merge_cells('A1:A3')
	sheetD['A4'].value='Reference'
	sheetD['B4'].value='Comment'
	sheetE.insert_cols(2)
	sheetE.move_range("A1:A3", cols=1)
	sheetE.merge_cells('A1:A3')
	sheetE['A4'].value='Reference'
	sheetE['B4'].value='Comment'
	for rows in sheetA.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetB.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetC.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetD.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetE.iter_rows():
		for cell in rows:
			cell.font = Font(name='Arial', size=11)
			cell.border = Border(left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), top=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.number_format = "0.00"
	for rows in sheetA.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
	for rows in sheetB.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
	for rows in sheetC.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
			cell.number_format = "0.000"
	for rows in sheetD.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
			cell.number_format = "0.000"
	for rows in sheetE.iter_rows(min_row=1, max_row=4, min_col=None):
		for cell in rows:
			cell.fill = PatternFill(start_color="000099", end_color="000099",fill_type = "solid")
			cell.font = Font(name='Arial', size=11, color='FFFFFF', bold=True)
	sheetA.column_dimensions['A'].width = 30
	sheetA.column_dimensions['B'].width = 12
	sheetB.column_dimensions['A'].width = 30
	sheetB.column_dimensions['B'].width = 12
	sheetC.column_dimensions['A'].width = 30
	sheetC.column_dimensions['B'].width = 12
	sheetD.column_dimensions['A'].width = 30
	sheetD.column_dimensions['B'].width = 12
	sheetE.column_dimensions['A'].width = 30
	sheetE.column_dimensions['B'].width = 12
	wb_style_prod.save(outputName)
	print('\n')
	print('Opening Trended Data in Excel, this may take a few moments.')

	
for file in os.listdir(wd):
	#print(file)
	if file.startswith('~$'):
		continue
	elif file.endswith('.xlsx'):
		trendedData = trendHPLC(wd,file)